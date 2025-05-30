from openpyxl import load_workbook
from geopy.geocoders import Nominatim
from time import sleep


def criar_geolocalizador():
    return Nominatim(user_agent='automacao_geolocalizacao')


def buscar_coordenadas(geolocalizador, endereco):
    try:
        local = geolocalizador.geocode(f"{endereco}, Blumenau")
        if local:
            return local.latitude, local.longitude
    except:
        pass
    return None, None


def preencher_coordenadas(arquivo):
    planilha = load_workbook(arquivo)
    manipular_planilha = planilha.active
    geo = criar_geolocalizador()

    for i, linha in enumerate(manipular_planilha.iter_rows(min_row=2, max_row=444, min_col=4, max_col=4, values_only=True), start=2):
        endereco = linha[0]

        if endereco is None:
            print(f"Linha {i}: sem endereço")
            manipular_planilha[f'E{i}'] = ''
            manipular_planilha[f'F{i}'] = ''
        else:
            latitude, longitude = buscar_coordenadas(geo, endereco)
            if latitude and longitude:
                print(f"Linha {i}: coordenadas OK")
                manipular_planilha[f'E{i}'] = latitude
                manipular_planilha[f'F{i}'] = longitude
            else:
                print(f"Linha {i}: não achou")
                manipular_planilha[f'E{i}'] = ''
                manipular_planilha[f'F{i}'] = ''

        sleep(1)

        planilha.save('obras_disponiveis_com_coordenadas.xlsx')
        print("Pronto! Salvo com coordenadas.")


preencher_coordenadas('obras_disponiveis.xlsx')